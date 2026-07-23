"""
가격 비교표 자동 생성 프로토타입.

우리 품목 리스트 → 품목마다 나라장터 유사 후보 top-N → 스샷과 같은 레이아웃의 엑셀 1장.
후보 상품 이미지는 나라장터에서 직접 다운로드해 셀에 임베드(수작업 캡처 불필요).

    .venv/bin/python build_report.py --out 비교표.xlsx --top 3

우리 품목은 지금 스샷의 지피코리아 예시를 시드로 넣어둠(SEED). 실제 엑셀 받으면 SEED만 교체.
"""
import argparse
import io
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor

import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import matcher

# ── 우리 품목 시드 (스샷의 지피코리아). 실제 엑셀 주면 이 리스트만 갈아끼움 ──
OUR_COMPANY = "주식회사 지피코리아"
SEED = [
    {"품명": "안내판", "모델": "GK-BF02", "규격": (500, 450), "재질": "스테인리스",
     "용도": "시각장애인용점자안내", "가격": 1_300_000, "식별번호": "26187495",
     "keyword": "점자 안내판"},
    {"품명": "안내판", "모델": "GK-BF05", "규격": (500, 450), "재질": "스테인리스",
     "용도": "시각장애인용점자/공원안내", "가격": 5_200_000, "식별번호": "26224093",
     "keyword": "점자 안내판"},
]

ROWS = ["물품식별번호", "규 격", "모 델 명", "재 질", "용 도", "가 격", "사 진"]

HEADER_FILL = PatternFill("solid", fgColor="D9D9F3")
OUR_FILL = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="9AA0B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
BOLD = Font(bold=True)


def _fmt(our, key):
    """우리 품목 dict → 구분 라벨에 해당하는 표시값."""
    if key == "물품식별번호":
        return our.get("식별번호", "")
    if key == "규 격":
        d = our.get("규격")
        return f"{d[0]}×{d[1]}mm" if d else ""
    if key == "모 델 명":
        return our.get("모델", "")
    if key == "재 질":
        return our.get("재질", "")
    if key == "용 도":
        return our.get("용도", "")
    if key == "가 격":
        return f"{our.get('가격', 0):,}"
    return ""


def _fmt_cand(c, key):
    if key == "물품식별번호":
        return c.get("식별번호", "")
    if key == "규 격":
        d = c.get("규격")
        return f"{d[0]}×{d[1]}mm" if d else ""
    if key == "모 델 명":
        return c.get("모델", "")
    if key == "재 질":
        return c.get("재질", "")
    if key == "용 도":
        return c.get("용도", "")
    if key == "가 격":
        return f"{c.get('가격', 0):,}"
    return ""


def _put(ws, r, col, val, fill=None, bold=False):
    cell = ws.cell(row=r, column=col, value=val)
    cell.alignment = CENTER
    cell.border = BORDER
    if fill:
        cell.fill = fill
    if bold:
        cell.font = BOLD
    return cell


def _add_image(ws, r, col, url, session, tmpdir):
    """url 이미지를 내려받아 (r,col) 셀에 임베드. 실패하면 조용히 건너뜀."""
    if not url:
        return
    try:
        full = url if url.startswith("http") else "https://shop.g2b.go.kr" + url
        data = session.get(full, timeout=30).content
        img = XLImage(io.BytesIO(data))
        img.width, img.height = 110, 100
        ws.add_image(img, f"{get_column_letter(col)}{r}")
    except Exception as e:
        _put(ws, r, col, f"(이미지 실패)")


def _match_one(our, top_n, max_pool, exact_price, excl, fallback_band):
    return our, matcher.find_candidates(
        our, keyword=our.get("keyword"), top_n=top_n, max_pool=max_pool,
        exact_price=exact_price, exclude_companies=excl, fallback_band=fallback_band)


def build(our_items, out_path, top_n=3, our_company=OUR_COMPANY, exclude_companies=None,
          max_pool=200, exact_price=False, fallback_band=0.0):
    excl = exclude_companies if exclude_companies is not None else [our_company]
    pairs = [_match_one(it, top_n, max_pool, exact_price, excl, fallback_band) for it in our_items]
    wb = Workbook(); ws = wb.active; ws.title = "가격비교표"
    session = requests.Session(); tmpdir = tempfile.mkdtemp()
    _render_pairs(ws, our_company, pairs, top_n, session, tmpdir)
    wb.save(out_path)
    return out_path


def build_multi(companies, out_path, top_n=3, max_pool=200, exact_price=True,
                exclude_companies=None, fallback_band=0.0, workers=8):
    """{업체명: [our_item...]} → 업체별 시트. 동시(ThreadPool) 매칭 후 렌더. 기본 계열사 상호 제외."""
    excl = exclude_companies if exclude_companies is not None else list(companies)
    tasks = [(name, i, it) for name, items in companies.items() for i, it in enumerate(items)]

    def work(t):
        name, i, it = t
        _, cands = _match_one(it, top_n, max_pool, exact_price, excl, fallback_band)
        mark = " (근사)" if cands and cands[0].get("근사") else ""
        print(f"[{name}] {it.get('모델') or it.get('품명')} ({it.get('가격') or 0:,}) → 후보 {len(cands)}개{mark}")
        return name, i, it, cands

    results = {}
    with ThreadPoolExecutor(max_workers=workers) as ex:   # 품목 단위 동시 매칭
        for name, i, it, cands in ex.map(work, tasks):
            results.setdefault(name, {})[i] = (it, cands)

    wb = Workbook(); wb.remove(wb.active)
    session = requests.Session(); tmpdir = tempfile.mkdtemp()
    for name, items in companies.items():
        ws = wb.create_sheet(name[:31])
        _render_pairs(ws, name, [results[name][i] for i in range(len(items))], top_n, session, tmpdir)
    wb.save(out_path)
    return out_path


def _render_pairs(ws, our_company, pairs, top_n, session, tmpdir):
    ncol = 3 + 1 + top_n                      # 연번/품명/구분 + 우리 + 후보N

    # 제목/업체명
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ncol)
    t = _put(ws, 1, 1, "가격 비교표", bold=True)
    t.font = Font(bold=True, size=16)
    _put(ws, 2, 1, f"○ 업체명 : {our_company}")

    r = 3
    for i, (our, cands) in enumerate(pairs, 1):
        # 블록 헤더 (연번/품명/구분/회사명들). band 폴백 후보는 (근사) 표기.
        _put(ws, r, 1, "연번", HEADER_FILL, bold=True)
        _put(ws, r, 2, "품 명", HEADER_FILL, bold=True)
        _put(ws, r, 3, "구분", HEADER_FILL, bold=True)
        _put(ws, r, 4, our_company, OUR_FILL, bold=True)
        for j, c in enumerate(cands):
            _put(ws, r, 5 + j, c["업체"] + (" (근사)" if c.get("근사") else ""), HEADER_FILL, bold=True)
        for j in range(len(cands), top_n):
            _put(ws, r, 5 + j, "", HEADER_FILL)

        # 상세 7행
        for k, key in enumerate(ROWS):
            rr = r + 1 + k
            _put(ws, rr, 3, key)                              # 구분 라벨
            if key == "사 진":
                ws.row_dimensions[rr].height = 80
                # 우리 이미지는 없음(추후 추가). 후보 이미지 임베드.
                for j, c in enumerate(cands):
                    _put(ws, rr, 5 + j, "")
                    _add_image(ws, rr, 5 + j, c.get("imgSrc"), session, tmpdir)
                _put(ws, rr, 4, "")
            else:
                _put(ws, rr, 4, _fmt(our, key), bold=(key == "가 격"))
                for j, c in enumerate(cands):
                    _put(ws, rr, 5 + j, _fmt_cand(c, key), bold=(key == "가 격"))
                for j in range(len(cands), top_n):
                    _put(ws, rr, 5 + j, "")

        # 연번/품명 세로 병합 (상세 7행)
        ws.merge_cells(start_row=r + 1, end_row=r + 7, start_column=1, end_column=1)
        ws.merge_cells(start_row=r + 1, end_row=r + 7, start_column=2, end_column=2)
        _put(ws, r + 1, 1, i, bold=True)
        _put(ws, r + 1, 2, our["품명"], bold=True)

        r += 8

    # 열 너비
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    for col in range(4, ncol + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="비교표.xlsx")
    ap.add_argument("--top", type=int, default=3)
    args = ap.parse_args()
    path = build(SEED, args.out, top_n=args.top)
    print("저장:", os.path.abspath(path))


if __name__ == "__main__":
    main()
