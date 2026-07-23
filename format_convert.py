"""
나라장터 '품목등록결과'(.xls) → 협상품목리스트(분석용)(.xlsx) 변환 CLI.

입력 파일만 주면 각 시트(업체)의 헤더를 **헤더명 기준으로 자동인식**해서 따라감.
시트마다 컬럼 위치가 달라도(예: 지피가든 22열) 안전. 하드코딩 위치 없음.

    python format_convert.py "입력.xls"
    python format_convert.py "입력.xls" --out 협상품목리스트.xlsx --days 40 --qty 5

매핑:  식별번호←물품식별번호 · 품명←한글품명 첫토큰 · 규격←한글품명 · 희망가←희망가(부가세포함)
       납품일수/공급예정수량 = 고정값(--days/--qty).  식별번호 없으면 빈칸으로 둠(한글품명 기준).
"""
import argparse

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT_HDR = ["식별번호", "품 명", "규 격", "납품일수", "공급예정수량", "희망가\n(부가세포함)"]
WIDTHS = [14, 8, 52, 10, 13, 15]

THIN = Side(style="thin", color="9AA0B4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
HDR_FILL = PatternFill("solid", fgColor="D9E1F2")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GOLD = PatternFill("solid", fgColor="FFD966")
BOLD = Font(bold=True)


def _norm(h):
    return str(h).replace("\n", "").replace("*", "").strip()


def _find_header(sh):
    """헤더 행 + {정규화헤더명: 열인덱스}. 물품식별번호/한글품명 있는 행을 헤더로."""
    for r in range(min(sh.nrows, 15)):
        cells = [_norm(sh.cell_value(r, c)) for c in range(sh.ncols)]
        if "물품식별번호" in cells and "한글품명" in cells:
            return r, {c: i for i, c in enumerate(cells) if c}
    raise ValueError("헤더(물품식별번호/한글품명)를 못 찾음")


def _clean_id(v):
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def _to_int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return v


def read_rows(sh):
    """시트 → [[식별번호, 품명, 규격, None, None, 희망가], ...] (헤더명 자동매핑)."""
    hr, col = _find_header(sh)
    ci_id, ci_nm, ci_pr = col["물품식별번호"], col["한글품명"], col.get("희망가(부가세포함)")
    out = []
    for r in range(hr + 1, sh.nrows):
        spec = str(sh.cell_value(r, ci_nm)).strip()   # 한글품명 = 규격(패킹 문자열), 앞 탭 제거
        if not spec:
            continue                                   # 한글품명이 실질 키
        idnf = _clean_id(sh.cell_value(r, ci_id))
        pumnm = spec.split(",")[0].strip() or "품목"
        price = _to_int(sh.cell_value(r, ci_pr)) if ci_pr is not None else ""
        out.append([idnf, pumnm, spec, None, None, price])
    return out


def _write_sheet(ws, company, rows, days, qty):
    cat = rows[0][1] if rows else "품목"
    ws.merge_cells("A1:F1")
    t = ws["A1"]; t.value = f"{company} '{cat}' 협상품목리스트(분석용)"
    t.font = Font(bold=True, size=14); t.alignment = LEFT
    ws.append([])
    ws.append(OUT_HDR)
    for c in range(1, 7):
        cell = ws.cell(row=3, column=c)
        cell.font = BOLD; cell.alignment = CENTER; cell.border = BORDER
        cell.fill = GOLD if c == 6 else HDR_FILL
    for row in rows:
        row[3], row[4] = days, qty
        ws.append(row)
        rr = ws.max_row
        for c in range(1, 7):
            cell = ws.cell(row=rr, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c == 3 else CENTER
            if c == 5:
                cell.fill = YELLOW
            if c == 6:
                cell.number_format = "#,##0"
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[chr(64 + i)].width = w


def read_all(in_path):
    wb_in = xlrd.open_workbook(in_path)
    return {name: read_rows(wb_in.sheet_by_name(name)) for name in wb_in.sheet_names()}


def to_markdown(rows, days, qty, max_spec=48):
    """행 목록 → 마크다운 표 문자열. 규격이 길면 잘라서 표시."""
    cols = ["식별번호", "품명", "규격", "납품일수", "공급예정수량", "희망가(부가세포함)"]
    md = ["| " + " | ".join(cols) + " |", "|" + "|".join([":--"] * len(cols)) + "|"]
    for idnf, pumnm, spec, _, _, price in rows:
        s = str(spec).replace("|", "/")
        if len(s) > max_spec:
            s = s[:max_spec - 1] + "…"
        pr = f"{price:,}" if isinstance(price, int) else price
        md.append(f"| {idnf or '-'} | {pumnm} | {s} | {days} | {qty} | {pr} |")
    return "\n".join(md)


def write_xlsx(data, out_path, days, qty):
    wb = Workbook(); wb.remove(wb.active)
    for name, rows in data.items():
        _write_sheet(wb.create_sheet(name[:31]), name, rows, days, qty)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="품목등록결과 .xls 경로")
    ap.add_argument("--out", default="협상품목리스트_분석용.xlsx")
    ap.add_argument("--days", type=int, default=40, help="납품일수 고정값")
    ap.add_argument("--qty", type=int, default=5, help="공급예정수량 고정값")
    ap.add_argument("--rows", type=int, default=0, help="시트당 미리보기 행수(0=전체)")
    ap.add_argument("--no-xlsx", action="store_true", help="미리보기만, 엑셀 저장 안 함")
    a = ap.parse_args()

    data = read_all(a.input)
    for name, rows in data.items():                       # 엑셀 전에 마크다운 표로 미리보기
        shown = rows[:a.rows] if a.rows else rows
        print(f"\n## {name}  ({len(rows)}건)\n")
        print(to_markdown(shown, a.days, a.qty))
        if a.rows and len(rows) > a.rows:
            print(f"\n_… 외 {len(rows) - a.rows}건_")

    total = sum(len(r) for r in data.values())
    if a.no_xlsx:
        print(f"\n미리보기 전용 (총 {total}건, 엑셀 미저장)")
    else:
        write_xlsx(data, a.out, a.days, a.qty)
        print(f"\n엑셀 저장 → {a.out}  (총 {total}건)")


if __name__ == "__main__":
    main()
